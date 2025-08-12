# Module holding all the extractors for the different cfdi types

# Importing required modules
import xml.etree.ElementTree as ET
from openpyxl import Workbook, load_workbook

# Parse the XML file

# Function to strip namespaces
def strip_namespace(tag):
    if '}' in tag:
        return tag.split('}', 1)[1]  # Remove namespace
    return tag

# Helper function to find a tag regardless of namespace
def find_all_tags(root, tag_name):
    return [elem for elem in root.iter() if strip_namespace(elem.tag) == tag_name]

# Function to parse "Pago" CFDI type
def parse_P(xml_file):
    namespaces = {
        'cfdi': 'http://www.sat.gob.mx/cfd/4',
        'pago20': 'http://www.sat.gob.mx/Pagos20',
        'tfd': 'http://www.sat.gob.mx/TimbreFiscalDigital'
    }

    # Parsear el archivo XML
    tree = ET.parse(xml_file)
    root = tree.getroot()

    # Extraer datos del emisor, receptor, timbre y complementos
    emisor = root.find('cfdi:Emisor', namespaces).attrib
    receptor = root.find('cfdi:Receptor', namespaces).attrib
    timbre = root.find('cfdi:Complemento/tfd:TimbreFiscalDigital', namespaces).attrib

    pagos = []
    for pago in root.findall('cfdi:Complemento/pago20:Pagos/pago20:Pago', namespaces):
        doctos_relacionados = []
        for docto in pago.findall('pago20:DoctoRelacionado', namespaces):
            doctos_relacionados.append({
                'IdDocumento': docto.attrib.get('IdDocumento'),
                'Serie': docto.attrib.get('Serie'),
                'Folio': docto.attrib.get('Folio'),
                'MonedaDR': docto.attrib.get('MonedaDR'),
                'EquivalenciaDR': docto.attrib.get('EquivalenciaDR'),
                'NumParcialidad': docto.attrib.get('NumParcialidad'),
                'ImpSaldoAnt': docto.attrib.get('ImpSaldoAnt'),
                'ImpPagado': docto.attrib.get('ImpPagado'),
                'ImpSaldoInsoluto': docto.attrib.get('ImpSaldoInsoluto'),
                'ObjetoImpDR': docto.attrib.get('ObjetoImpDR')
            })
        pagos.append({
            'FechaPago': pago.attrib.get('FechaPago'),
            'FormaDePagoP': pago.attrib.get('FormaDePagoP'),
            'MonedaP': pago.attrib.get('MonedaP'),
            'TipoCambioP': pago.attrib.get('TipoCambioP'),
            'Monto': pago.attrib.get('Monto'),
            'DoctosRelacionados': doctos_relacionados
        })

    return {
        'Emisor': emisor,
        'Receptor': receptor,
        'TimbreFiscal': timbre,
        'Pagos': pagos
    }

#function to write "pago" CFDI type to excel
def writeP_to_excel(data, output_file):
    # Intentar cargar el archivo Excel, o crear uno nuevo
    try:
        wb = load_workbook(output_file)
    except FileNotFoundError:
        wb = Workbook()

    # Verificar si existe la hoja "Pagos"
    if "Pagos" not in wb.sheetnames:
        sheet = wb.create_sheet("Pagos")
        # Escribir encabezados
        sheet.append([
            "UUID Timbre", "Fecha Timbrado",
            "RFC Emisor", "Nombre Emisor", "Régimen Fiscal Emisor",
            "RFC Receptor", "Nombre Receptor",
            "Fecha Pago", "Forma De Pago P", "Moneda P", "Tipo Cambio P", "Monto",
            "Id Documento", "Serie", "Folio", "Moneda DR", "Equivalencia DR",
            "Num Parcialidad", "Imp Saldo Ant", "Imp Pagado", "Imp Saldo Insoluto", "Objeto Imp DR"
        ])
    else:
        sheet = wb["Pagos"]

    # Agregar los datos al Excel
    for pago in data['Pagos']:
        for docto in pago['DoctosRelacionados']:
            sheet.append([
                data['TimbreFiscal']['UUID'], data['TimbreFiscal']['FechaTimbrado'],
                data['Emisor'].get('Rfc'), data['Emisor'].get('Nombre'), data['Emisor'].get('RegimenFiscal'),
                data['Receptor'].get('Rfc'), data['Receptor'].get('Nombre'),
                pago.get('FechaPago'), pago.get('FormaDePagoP'), pago.get('MonedaP'),
                pago.get('TipoCambioP'), pago.get('Monto'),
                docto.get('IdDocumento'), docto.get('Serie'), docto.get('Folio'),
                docto.get('MonedaDR'), docto.get('EquivalenciaDR'),
                docto.get('NumParcialidad'), docto.get('ImpSaldoAnt'),
                docto.get('ImpPagado'), docto.get('ImpSaldoInsoluto'), docto.get('ObjetoImpDR')
            ])

    # Guardar el archivo Excel
    wb.save(output_file)

# Function to parse "Ingreso and Egreso" CFDI type
# Función para guardar los datos en Excel con el orden solicitado
def saveIE_to_excel(data, output_file):
    # Determinar el nombre de la hoja basado en el tipo de comprobante
    sheet_name = "Ingresos" if data['Comprobante']['TipoDeComprobante'] == 'I' else "Egresos"

    # Intentar cargar el archivo Excel si existe, de lo contrario, crear uno nuevo
    try:
        wb = load_workbook(output_file)
    except FileNotFoundError:
        wb = Workbook()

    # Verificar si existe la hoja correspondiente, de lo contrario, crearla
    if sheet_name not in wb.sheetnames:
        sheet = wb.create_sheet(sheet_name)
        # Escribir encabezados
        sheet.append([
            "UUID", "Fecha", "Serie", "Folio", "Tipo", 
            "RFC Emisor", "Nombre Emisor", "Régimen Fiscal",
            "Cantidad", "Valor Unitario", "Importe", "Traslado Importe",
            "Subtotal", "Total", "Forma de Pago", "Descripción", 
            "Moneda", "Uso CFDI", 
            "RFC Receptor", "Nombre Receptor", "Domicilio", "Régimen Fiscal",
            "Traslado Base", "Fecha Timbrado", "Versión"
        ])
    else:
        sheet = wb[sheet_name]

    # Agregar datos
    for concepto in data['Conceptos']:
        sheet.append([
            data['TimbreFiscal']['UUID'], data['Comprobante']['Fecha'], 
            data['Comprobante']['Serie'], data['Comprobante']['Folio'], 
            data['Comprobante']['TipoDeComprobante'],
            data['Emisor'].get('Rfc'), data['Emisor'].get('Nombre'), data['Emisor'].get('RegimenFiscal'),
            concepto.get('Cantidad'), concepto.get('ValorUnitario'), concepto.get('Importe'), concepto.get('Traslado_Importe'),
            data['Comprobante']['SubTotal'], data['Comprobante']['Total'], data['Comprobante']['FormaPago'],
            concepto.get('Descripcion'), data['Comprobante']['Moneda'], 
            data['Receptor'].get('UsoCFDI'),
            data['Receptor'].get('Rfc'), data['Receptor'].get('Nombre'), data['Receptor'].get('DomicilioFiscalReceptor'),
            data['Receptor'].get('RegimenFiscalReceptor'), concepto.get('Traslado_Base'),
            data['TimbreFiscal']['FechaTimbrado'], data['Comprobante']['Version']
        ])

    # Guardar el archivo Excel
    wb.save(output_file)

# Función para parsear el XML (sin cambios adicionales ya que la extracción parece correcta)
def parse_IE(xml_file):
    namespaces = {
        'cfdi': 'http://www.sat.gob.mx/cfd/4',
        'tfd': 'http://www.sat.gob.mx/TimbreFiscalDigital'
    }

    tree = ET.parse(xml_file)
    root = tree.getroot()

    comprobante = {
        'Version': root.attrib.get('Version'),
        'Serie': root.attrib.get('Serie'),
        'Folio': root.attrib.get('Folio'),
        'Fecha': root.attrib.get('Fecha'),
        'SubTotal': root.attrib.get('SubTotal'),
        'Total': root.attrib.get('Total'),
        'FormaPago': root.attrib.get('FormaPago'),
        'TipoDeComprobante': root.attrib.get('TipoDeComprobante'),
        'Moneda': root.attrib.get('Moneda')
    }

    emisor = root.find('cfdi:Emisor', namespaces).attrib
    receptor = root.find('cfdi:Receptor', namespaces).attrib

    conceptos = []
    for concepto in root.findall('cfdi:Conceptos/cfdi:Concepto', namespaces):
        concepto_data = {
            'Descripcion': concepto.attrib.get('Descripcion'),
            'Cantidad': concepto.attrib.get('Cantidad'),
            'ValorUnitario': concepto.attrib.get('ValorUnitario'),
            'Importe': concepto.attrib.get('Importe'),
        }

        traslado = concepto.find('cfdi:Impuestos/cfdi:Traslados/cfdi:Traslado', namespaces)
        if traslado is not None:
            concepto_data['Traslado_Base'] = traslado.attrib.get('Base')
            concepto_data['Traslado_Importe'] = traslado.attrib.get('Importe')

        conceptos.append(concepto_data)

    complemento = root.find('cfdi:Complemento/tfd:TimbreFiscalDigital', namespaces)
    timbre_fiscal = {
        'UUID': complemento.attrib.get('UUID'),
        'FechaTimbrado': complemento.attrib.get('FechaTimbrado')
    }

    return {
        'Comprobante': comprobante,
        'Emisor': emisor,
        'Receptor': receptor,
        'Conceptos': conceptos,
        'TimbreFiscal': timbre_fiscal
    }

    # Guardar el archivo Excel
    wb.save(output_file)
    # Function to write "pago" CFDI type to excel

# Function to parse "Nomina" CFDI type
def parse_N(xml_file):
    namespaces = {
        'cfdi': 'http://www.sat.gob.mx/cfd/4',
        'tfd': 'http://www.sat.gob.mx/TimbreFiscalDigital',
        'nomina12': 'http://www.sat.gob.mx/nomina12'
    }

    # Parsear el archivo XML
    tree = ET.parse(xml_file)
    root = tree.getroot()

    # Extraer información del comprobante
    comprobante = {
        'Serie': root.attrib.get('Serie', 'N/A'),
        'Folio': root.attrib.get('Folio', 'N/A'),
        'Fecha': root.attrib.get('Fecha', 'N/A'),
        'Moneda': root.attrib.get('Moneda', 'N/A'),
        'SubTotal': root.attrib.get('SubTotal', '0.00'),
        'Descuento': root.attrib.get('Descuento', '0.00'),
        'Total': root.attrib.get('Total', '0.00'),
    }

    # Extraer información del timbre fiscal
    complemento = root.find('cfdi:Complemento/tfd:TimbreFiscalDigital', namespaces)
    timbre_fiscal = {
        'UUID': complemento.attrib.get('UUID', 'N/A'),
        'FechaTimbrado': complemento.attrib.get('FechaTimbrado', 'N/A')
    }

    # Extraer información del emisor
    emisor = root.find('cfdi:Emisor', namespaces).attrib

    # Extraer información del receptor
    receptor = root.find('cfdi:Receptor', namespaces).attrib

    # Extraer conceptos
    conceptos = []
    for concepto in root.findall('cfdi:Conceptos/cfdi:Concepto', namespaces):
        concepto_data = {
            'Descripcion': concepto.attrib.get('Descripcion', 'N/A'),
            'Cantidad': concepto.attrib.get('Cantidad', '0'),
            'ValorUnitario': concepto.attrib.get('ValorUnitario', '0.00'),
            'Importe': concepto.attrib.get('Importe', '0.00'),
        }
        conceptos.append(concepto_data)

    # Extraer datos del complemento de nómina
    complemento_nomina = root.find('cfdi:Complemento/nomina12:Nomina', namespaces)
    nomina = {
        'Version': complemento_nomina.attrib.get('Version', 'N/A') if complemento_nomina is not None else 'N/A',
        'TipoNomina': complemento_nomina.attrib.get('TipoNomina', 'N/A') if complemento_nomina is not None else 'N/A',
        'TotalPercepciones': complemento_nomina.attrib.get('TotalPercepciones', '0.00') if complemento_nomina is not None else '0.00',
        'TotalDeducciones': complemento_nomina.attrib.get('TotalDeducciones', '0.00') if complemento_nomina is not None else '0.00',
        'TotalOtrosPagos': complemento_nomina.attrib.get('TotalOtrosPagos', '0.00') if complemento_nomina is not None else '0.00'
    }

    # Extraer percepciones
    percepciones = []
    if complemento_nomina is not None:
        percepciones_elem = complemento_nomina.find('nomina12:Percepciones', namespaces)
        if percepciones_elem is not None:
            for percepcion in percepciones_elem.findall('nomina12:Percepcion', namespaces):
                percepciones.append({
                    'Clave': percepcion.attrib.get('Clave', 'N/A'),
                    'Concepto': percepcion.attrib.get('Concepto', 'N/A'),
                    'ImporteGravado': percepcion.attrib.get('ImporteGravado', '0.00'),
                    'ImporteExento': percepcion.attrib.get('ImporteExento', '0.00')
                })

    # Extraer deducciones
    deducciones = []
    if complemento_nomina is not None:
        deducciones_elem = complemento_nomina.find('nomina12:Deducciones', namespaces)
        if deducciones_elem is not None:
            for deduccion in deducciones_elem.findall('nomina12:Deduccion', namespaces):
                deducciones.append({
                    'Clave': deduccion.attrib.get('Clave', 'N/A'),
                    'Concepto': deduccion.attrib.get('Concepto', 'N/A'),
                    'Importe': deduccion.attrib.get('Importe', '0.00')
                })

    # Extraer otros pagos
    otros_pagos = []
    if complemento_nomina is not None:
        otros_pagos_elem = complemento_nomina.find('nomina12:OtrosPagos', namespaces)
        if otros_pagos_elem is not None:
            for otro_pago in otros_pagos_elem.findall('nomina12:OtroPago', namespaces):
                otros_pagos.append({
                    'Clave': otro_pago.attrib.get('Clave', 'N/A'),
                    'Concepto': otro_pago.attrib.get('Concepto', 'N/A'),
                    'Importe': otro_pago.attrib.get('Importe', '0.00')
                })

    return {
        'TimbreFiscal': timbre_fiscal,
        'Comprobante': comprobante,
        'Emisor': emisor,
        'Receptor': receptor,
        'Conceptos': conceptos,
        'Nomina': nomina,
        'Percepciones': percepciones,
        'Deducciones': deducciones,
        'OtrosPagos': otros_pagos
    }

# Function to save "Nomina" CFDI type to excel
def saveN_to_excel(data, output_file):
    sheet_name = "Nómina"

    try:
        wb = load_workbook(output_file)
    except FileNotFoundError:
        wb = Workbook()

    if sheet_name not in wb.sheetnames:
        sheet = wb.create_sheet(sheet_name)
        sheet.append([
            "UUID", "Fecha Timbrado",
            "Serie", "Folio", "Fecha", "Moneda", "SubTotal", "Descuento", "Total",
            "RFC Emisor", "Nombre Emisor",
            "RFC Receptor", "Nombre Receptor",
            "Descripcion", "Cantidad", "Valor Unitario", "Importe",
            "Version Nómina", "Tipo Nómina", "Total Percepciones", "Total Deducciones", "Total Otros Pagos"
        ])
    else:
        sheet = wb[sheet_name]

    for concepto in data['Conceptos']:
        sheet.append([
            data['TimbreFiscal']['UUID'], data['TimbreFiscal']['FechaTimbrado'],
            data['Comprobante']['Serie'], data['Comprobante']['Folio'], data['Comprobante']['Fecha'],
            data['Comprobante']['Moneda'], data['Comprobante']['SubTotal'], data['Comprobante']['Descuento'], data['Comprobante']['Total'],
            data['Emisor'].get('Rfc', 'N/A'), data['Emisor'].get('Nombre', 'N/A'),
            data['Receptor'].get('Rfc', 'N/A'), data['Receptor'].get('Nombre', 'N/A'),
            concepto.get('Descripcion', 'N/A'), concepto.get('Cantidad', '0'), concepto.get('ValorUnitario', '0.00'), concepto.get('Importe', '0.00'),
            data['Nomina']['Version'], data['Nomina']['TipoNomina'], data['Nomina']['TotalPercepciones'], data['Nomina']['TotalDeducciones'], data['Nomina']['TotalOtrosPagos']
        ])

    wb.save(output_file)